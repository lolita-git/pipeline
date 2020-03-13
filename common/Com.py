import os
from xlrd import open_workbook
import openpyxl

#去掉写数据到excel
class Common(object):
    def __init__(self):
        self.MAIN_PATH = os.path.dirname(os.path.abspath(__file__)).strip("\\common")


    def getXls(self, xls_name, sheet_name):
        xlsPath = os.path.join(self.MAIN_PATH, "testdata", xls_name)
        xlsPath = xlsPath.replace("\\", "/")
        file = open_workbook(xlsPath)
        sheet = file.sheet_by_name(sheet_name)
        nrows = sheet.nrows
        if nrows > 1:
            # 获取第一列的内容，列表格式
            keys = sheet.row_values(0)
            # print(keys)
            listApiData = []
            # 获取每一行的内容，列表格式
            for col in range(1, nrows):
                if sheet.row_values(col) != "":
                    values = sheet.row_values(col)
                    # keys，values这两个列表一一对应来组合转换为字典
                    api_dict = dict(zip(keys, values))
                    # print(api_dict)
                    listApiData.append(api_dict)

            return listApiData
        else:
            print("表格无数据")
            return None

    def writeCell(self, xls_name, sheetName, content, coordinate=None, rowNo=None, colsNo=None):
        xlsPath = os.path.join(self.MAIN_PATH, "testdata", xls_name)
        xlsPath = xlsPath.replace("\\", "/")
        workbook = openpyxl.load_workbook(xlsPath)
        sheet = workbook.get_sheet_by_name(sheetName)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                workbook.save(xlsPath)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                workbook.save(xlsPath)
            except Exception as e:
                raise e
                # logger.exception("wrongError", e)
        else:
            raise Exception("Insufficient Coordinates of cell")

if __name__=="__main__":
    gets_xls =Common()
    gets_xls.writeCell(u"发票接口测试数据.xls", "addNew", "self.total", rowNo=1, colsNo=10)


